"""
🍵 Matchya — Resume Ranker (portfolio-ready)

Features:
- OpenAI-compatible LLMs (OpenAI, OpenRouter, LM Studio, custom) extract the main role, full name, contacts, and scores per
  criterion with concise reasoning — one request per resume for reliability.
- Robust composite scoring: 0.75×weighted percentiles + 0.25×coverage.
- Strong deduplication: same files/texts, same emails/phones, and near-duplicates by similarity threshold.
- Export to XLSX with styling, duplicate risk highlights, and similarity pairs for audits.
- JSONL checkpointing by SHA1 for safe resumability.
- Sources: uploaded files or direct cloud links to resume files (PDF/DOCX/TXT/MD/RTF/HTML).

Quick start:
    pip install streamlit pdfminer.six python-docx rapidfuzz pandas openpyxl pydantic tenacity openai requests beautifulsoup4 lxml

Run:
    streamlit run app_requests.py
"""

from __future__ import annotations
import io, os, re, json, hashlib, urllib.parse
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional

import numpy as np
import pandas as pd
import streamlit as st
from pydantic import BaseModel, Field
from rapidfuzz import fuzz
import requests
from bs4 import BeautifulSoup

# ---------- Parsers ----------
try:
    from pdfminer.high_level import extract_text as pdf_extract_text
except Exception:
    pdf_extract_text = None
try:
    import docx
except Exception:
    docx = None

ALLOWED_EXT = {".pdf", ".docx", ".txt", ".md", ".rtf"}
TOKEN_SPLIT = re.compile(r"[\W_]+", re.UNICODE)
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
PHONE_CAND_RE = re.compile(r"(?:\+?\d[\d\-\s()\./]{6,}\d)")
URL_RE = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)

CT_TO_EXT = {
    "application/pdf": ".pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "text/plain": ".txt",
    "text/rtf": ".rtf",
    "application/rtf": ".rtf",
    "text/markdown": ".md",
}
EXT_GUESS_FALLBACK = [".pdf", ".docx", ".txt", ".rtf", ".md"]
MAX_DOWNLOAD_MB = 25
REQ_TIMEOUT = 30

DEFAULT_MODEL_CHOICES = {
    "openai": ["gpt-4o-mini", "gpt-4o", "gpt-4.1-mini", "gpt-4.1"],
    "openrouter": ["openrouter/auto", "anthropic/claude-3.5-sonnet", "openai/gpt-4o-mini"],
    "lmstudio": ["lmstudio-community/gpt-4o-mini-gguf", "lmstudio-community/llama-3.1-8b-instruct"],
    "custom": ["gpt-4o-mini"],
}

# --- simple full-name heuristic (fallback when LLM is empty) ---
NAME_RE = re.compile(r"\b[A-Z][a-z]+ [A-Z][a-z]+(?: [A-Z][a-z]+)?\b")


@dataclass
class LLMSettings:
    provider: str
    api_key: str
    model: str
    base_url: str = ""
    headers: Optional[Dict[str, str]] = None


@dataclass
class RoleContext:
    description: str
    criteria: List[Criterion]


@dataclass
class ResumeArtifact:
    id: str
    file_hash: str
    text_hash: str
    name: str
    text: str
    url: str = ""

def guess_full_name(text: str) -> str:
    lines = [l.strip() for l in text.splitlines() if l.strip()][:30]
    for l in lines:
        if EMAIL_RE.search(l) or PHONE_CAND_RE.search(l) or "http" in l.lower():
            continue
        m = NAME_RE.search(l)
        if m:
            return m.group(0)
    return ""

# ---------- Models ----------
class Criterion(BaseModel):
    name: str
    weight: float = Field(1.0, ge=0)
    keywords: List[str] = Field(default_factory=list)

class LLMScores(BaseModel):
    scores: Dict[str, float]
    reasoning: Dict[str, str] = Field(default_factory=dict)

# ---------- Utils ----------
def read_file_text(filename: str, bytes_data: bytes) -> str:
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".pdf":
        if not pdf_extract_text:
            raise RuntimeError("pdfminer.six is missing. pip install pdfminer.six")
        with io.BytesIO(bytes_data) as bio:
            return pdf_extract_text(bio)
    if ext == ".docx":
        if not docx:
            raise RuntimeError("python-docx is missing. pip install python-docx")
        with io.BytesIO(bytes_data) as bio:
            d = docx.Document(bio)
            return "\n".join(p.text for p in d.paragraphs)
    if ext in {".txt", ".md", ".rtf"}:
        try:
            return bytes_data.decode("utf-8", errors="ignore")
        except Exception:
            return bytes_data.decode("latin-1", errors="ignore")
    raise ValueError(f"Unsupported format: {ext}")

def html_to_text(html_bytes: bytes, base_url: str = "") -> str:
    try:
        soup = BeautifulSoup(html_bytes, "lxml")
    except Exception:
        soup = BeautifulSoup(html_bytes, "html.parser")
    for t in soup(["script", "style", "noscript", "template"]):
        t.decompose()
    for t in soup.find_all(["nav", "footer", "aside"]):
        t.decompose()
    title = ""
    try:
        title = (soup.title.string or "").strip()
    except Exception:
        pass
    text = soup.get_text(separator="\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    header_bits = []
    if title:
        header_bits.append(title)
    if base_url:
        header_bits.append(base_url)
    return ("\n".join(header_bits + [text])).strip()

def sha1_bytes(b: bytes) -> str:
    return hashlib.sha1(b).hexdigest()

def sha1_text(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8", errors="ignore")).hexdigest()

def normalize_for_sim(t: str) -> str:
    t = re.sub(r"\s+", " ", t.lower())
    return re.sub(r"[^a-zа-я0-9 ]+", " ", t)

def normalize_phone_digits(s: str) -> str:
    return "".join(ch for ch in s if ch.isdigit() or ch == "+")

def best_phone(phones: List[str]) -> str:
    cleaned = []
    for p in phones:
        norm = normalize_phone_digits(p)
        digits = "".join(d for d in norm if d.isdigit())
        if 10 <= len(digits) <= 15:
            cleaned.append(norm)
    if not cleaned:
        return ""
    cleaned.sort(key=lambda x: (not x.startswith("+"), -len(x)))
    return cleaned[0]

def chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i+n]

def clamp(s: str, max_len: int) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip())
    return (s[:max_len-1] + "…") if len(s) > max_len else s

def normalize_url(u: str) -> str:
    u = (u or "").strip().strip('"').strip("'")
    if not u:
        return ""
    u = u.replace(" ", "%20")
    if not re.match(r"^https?://", u, flags=re.I):
        return ""
    try:
        pr = urllib.parse.urlparse(u)
        if pr.netloc.lower().endswith("drive.google.com"):
            qs = urllib.parse.parse_qs(pr.query)
            if "id" in qs and qs["id"]:
                file_id = qs["id"][0]
                return f"https://drive.google.com/uc?export=download&id={file_id}"
            m = re.search(r"/file/d/([^/]+)/", pr.path)
            if m:
                file_id = m.group(1)
                return f"https://drive.google.com/uc?export=download&id={file_id}"
    except Exception:
        pass
    return u

def guess_filename_from_headers(url: str, resp: requests.Response) -> str:
    cd = resp.headers.get("Content-Disposition") or resp.headers.get("content-disposition") or ""
    m = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^\";]+)"?', cd, flags=re.I)
    if m:
        name = m.group(1).strip().strip('"')
        return name
    path = urllib.parse.urlparse(url).path
    base = os.path.basename(path) or "download"
    return base


def _safe_get_json(url: str, headers: Optional[Dict[str, str]] = None) -> Dict:
    resp = requests.get(url, headers=headers or {}, timeout=REQ_TIMEOUT)
    resp.raise_for_status()
    return resp.json()


@st.cache_data(show_spinner=False)
def fetch_provider_models(provider: str, api_key: str = "", base_url: str = "") -> List[str]:
    """Fetch model list for the UI; fall back to defaults when discovery fails."""
    provider = (provider or "").strip().lower()
    if provider not in DEFAULT_MODEL_CHOICES:
        return DEFAULT_MODEL_CHOICES["openai"]

    try:
        if provider == "openai" and api_key.strip():
            from openai import OpenAI

            client = OpenAI(api_key=api_key.strip())
            models = client.models.list().data
            ids = sorted({m.id for m in models if getattr(m, "id", "")})
            if ids:
                return ids
        elif provider == "openrouter" and api_key.strip():
            data = _safe_get_json(
                "https://openrouter.ai/api/v1/models",
                headers={"Authorization": f"Bearer {api_key.strip()}"},
            )
            ids = sorted({m.get("id", "") for m in data.get("data", []) if m.get("id")})
            if ids:
                return ids
        elif provider == "lmstudio":
            # Support both /v1 and plain roots to be resilient to user-supplied base URLs
            base = (base_url or "http://localhost:1234").rstrip("/")
            primary = base if base.endswith("/v1") else base + "/v1"
            candidates = [primary + "/models"]
            if base != primary:
                candidates.append(base + "/models")  # fallback for pre-suffixed URLs

            headers = {"Authorization": f"Bearer {api_key.strip()}"} if api_key.strip() else None
            for url in candidates:
                try:
                    data = _safe_get_json(url, headers=headers)
                    ids = sorted({m.get("id", "") for m in data.get("data", []) if m.get("id")})
                    if ids:
                        return ids
                except Exception:
                    continue
            raise RuntimeError("LM Studio did not return model IDs")
    except Exception as exc:
        st.warning(f"Failed to fetch models from {provider}: {exc}")

    return DEFAULT_MODEL_CHOICES[provider]

def ensure_allowed_extension(filename: str, content_type: str) -> str:
    ext = os.path.splitext(filename)[1].lower()
    if ext in ALLOWED_EXT:
        return filename
    ct = (content_type or "").split(";")[0].strip().lower()
    if ct in CT_TO_EXT:
        return (filename + CT_TO_EXT[ct]) if ext == "" else (os.path.splitext(filename)[0] + CT_TO_EXT[ct])
    for g in EXT_GUESS_FALLBACK:
        if filename.lower().endswith(g):
            return filename
    return filename + ".pdf"

def is_content_type_supported(content_type: str) -> bool:
    ct = (content_type or "").split(";")[0].strip().lower()
    return (ct in CT_TO_EXT) or any(ct.endswith(x) for x in ["pdf", "rtf", "plain", "markdown", "msword", "officedocument.wordprocessingml.document"])

def stream_download(url: str) -> Tuple[str, bytes, str]:
    headers = {"User-Agent": "Mozilla/5.0 (ResumeRankerBot/1.0)"}
    with requests.get(url, headers=headers, timeout=REQ_TIMEOUT, stream=True, allow_redirects=True) as r:
        r.raise_for_status()
        total = 0
        chunks_list = []
        for chunk in r.iter_content(chunk_size=8192):
            if chunk:
                chunks_list.append(chunk)
                total += len(chunk)
                if total > MAX_DOWNLOAD_MB * 1024 * 1024:
                    raise RuntimeError(f"File too large (> {MAX_DOWNLOAD_MB}MB)")
        data = b"".join(chunks_list)
        fname = guess_filename_from_headers(url, r)
        fname = ensure_allowed_extension(fname, r.headers.get("Content-Type",""))
        return fname, data, r.headers.get("Content-Type","")


def collect_local_directory(dir_path: str, recursive: bool = False) -> List[Dict[str, object]]:
    """Scan a local server directory and return files in the expected structure."""
    dir_path = (dir_path or "").strip()
    if not dir_path:
        return []
    if not os.path.exists(dir_path) or not os.path.isdir(dir_path):
        raise FileNotFoundError(f"{dir_path} does not exist or is not a directory")

    items: List[Dict[str, object]] = []
    walker = os.walk(dir_path) if recursive else [(dir_path, [], os.listdir(dir_path))]
    for root, _, files in walker:
        for fname in files:
            ext = os.path.splitext(fname)[1].lower()
            if ext not in ALLOWED_EXT:
                continue
            full_path = os.path.join(root, fname)
            try:
                with open(full_path, "rb") as f:
                    b = f.read()
                items.append({"kind": "file", "name": fname, "bytes": b, "source_path": full_path})
            except Exception as exc:  # keep logs human-readable
                print(f"[dir-scan] {full_path}: {exc}")
    return items

# ---------- LLM ----------
class LLMClient:
    """Thin wrapper for OpenAI-compatible LLMs.

    Works with OpenAI, OpenRouter, LM Studio, or any OpenAI-style API. Keeps
    the payload explicit for easier debugging.
    """

    def __init__(self, api_key: str, model: str = "gpt-4o-mini", base_url: str = "", extra_headers: Optional[Dict[str,str]] = None):
        from openai import OpenAI

        client_opts = {"api_key": api_key}
        if base_url:
            client_opts["base_url"] = base_url
        if extra_headers:
            client_opts["default_headers"] = extra_headers
        self.client = OpenAI(**client_opts)
        self.model = model

    def score_and_extract_batch(
        self,
        resumes: List[Dict[str, str]],
        role_desc: str,
        criteria: List[Criterion],
    ) -> List[Dict[str, object]]:
        """
        Batch request (kept for reference; single mode is the default UI path).
        """
        system = (
            "You are an HR assistant. For EACH resume in the list:\n"
            "1) Extract specialization: 'specialization_main' (<=80 chars) plus up to 3 alternatives.\n"
            "2) Extract candidate full name into 'full_name' (empty if unknown).\n"
            "3) Extract contacts: 'emails' and 'phones' (international format preferred).\n"
            "4) Score every criterion 0..5 with concise reasoning per criterion.\n"
            "Return JSON exactly matching the schema for all inputs."
        )
        payload = {
            "role_description": role_desc,
            "criteria": [c.model_dump() for c in criteria],
            "resumes": [{"id": r["id"], "text": r["text"][:18000]} for r in resumes][:5]
        }

        schema = {
            "type": "object",
            "properties": {
                "results": {
                    "type": "array",
                    "maxItems": 5,
                    "items": {
                        "type": "object",
                        "properties": {
                            "id": {"type": "string"},
                            "full_name": {"type": "string"},
                            "specialization_main": {"type": "string"},
                            "specialization_alt": {"type": "array", "items": {"type": "string"}},
                            "emails": {"type": "array", "items": {"type": "string"}},
                            "phones": {"type": "array", "items": {"type": "string"}},
                            "scores": {"type": "object", "additionalProperties": {"type": "number"}},
                            "reasoning": {"type": "object", "additionalProperties": {"type": "string"}}
                        },
                        "required": ["id","specialization_main","emails","phones","scores"],
                        "additionalProperties": False
                    }
                }
            },
            "required": ["results"],
            "additionalProperties": False
        }

        resp = self.client.chat.completions.create(
            model=self.model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": json.dumps(payload, ensure_ascii=False)}
            ],
            response_format={"type": "json_schema", "json_schema": {"name": "ResumeBatch", "schema": schema}},
            temperature=0.1,
        )
        content = resp.choices[0].message.content
        data = json.loads(content)
        return data.get("results", [])

    def score_and_extract_single(
        self,
        resume: Dict[str, str],
        role_desc: str,
        criteria: List[Criterion],
    ) -> Dict[str, object]:
        """Safer single-resume scoring without batching."""

        system = (
            "You are an HR assistant. You receive ONE resume.\n"
            "1) Extract specialization: 'specialization_main' (<=80 chars) plus up to 3 alternatives.\n"
            "2) Extract the candidate full name into 'full_name' (leave empty if unknown).\n"
            "3) Extract contacts: 'emails' and 'phones' (international format preferred).\n"
            "4) Score every criterion 0..5 with concise reasoning per criterion.\n"
            "Return JSON exactly following the schema."
        )

        payload = {
            "role_description": role_desc,
            "criteria": [c.model_dump() for c in criteria],
            "resume": {"id": resume["id"], "text": resume["text"][:18000]},
        }

        schema = {
            "type": "object",
            "properties": {
                "id": {"type": "string"},
                "full_name": {"type": "string"},
                "specialization_main": {"type": "string"},
                "specialization_alt": {"type": "array", "items": {"type": "string"}},
                "emails": {"type": "array", "items": {"type": "string"}},
                "phones": {"type": "array", "items": {"type": "string"}},
                "scores": {"type": "object", "additionalProperties": {"type": "number"}},
                "reasoning": {"type": "object", "additionalProperties": {"type": "string"}},
            },
            "required": ["id", "specialization_main", "emails", "phones", "scores"],
            "additionalProperties": False,
        }

        resp = self.client.chat.completions.create(
            model=self.model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
            ],
            response_format={"type": "json_schema", "json_schema": {"name": "ResumeSingle", "schema": schema}},
            temperature=0.1,
        )

        content = resp.choices[0].message.content
        data = json.loads(content)
        return data

    def suggest_criteria(self, role_desc: str, max_items: int = 10) -> List[Criterion]:
        """Generate skill/criteria suggestions from the role description."""
        prompt = {
            "role_description": role_desc,
            "format": "json",
            "max_items": max_items,
        }
        resp = self.client.chat.completions.create(
            model=self.model,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a career expert. Based on the job description, return up to 10 key criteria/skills. "
                        "Each item: name, weight (0.5..3.0, higher for must-haves), keywords (3-6)."
                    ),
                },
                {"role": "user", "content": json.dumps(prompt, ensure_ascii=False)},
            ],
            response_format={
                "type": "json_schema",
                "json_schema": {
                    "name": "Criteria", "schema": {
                        "type": "object",
                        "properties": {
                            "criteria": {
                                "type": "array",
                                "maxItems": max_items,
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "name": {"type": "string"},
                                        "weight": {"type": "number"},
                                        "keywords": {"type": "array", "items": {"type": "string"}},
                                    },
                                    "required": ["name", "weight"],
                                    "additionalProperties": False,
                                },
                            }
                        },
                        "required": ["criteria"],
                        "additionalProperties": False,
                    }
                }
            },
            temperature=0.2,
        )
        raw = json.loads(resp.choices[0].message.content)
        out = []
        for item in raw.get("criteria", []):
            try:
                out.append(Criterion(**item))
            except Exception:
                continue
        return out


def create_llm_client(settings: LLMSettings) -> LLMClient:
    provider = (settings.provider or "openai").lower()
    if provider == "lmstudio":
        base_url = settings.base_url.strip() or "http://localhost:1234/v1"
        key = settings.api_key.strip() or "lm-studio"
        return LLMClient(api_key=key, model=settings.model, base_url=base_url)
    if provider == "openrouter":
        key = settings.api_key.strip()
        if not key:
            raise ValueError("OpenRouter API Key is required")
        headers = settings.headers or {
            "HTTP-Referer": "https://github.com/user/portfolio",  # good manners for OpenRouter
            "X-Title": "Resume ranker",
        }
        return LLMClient(api_key=key, model=settings.model, base_url="https://openrouter.ai/api/v1", extra_headers=headers)
    if provider == "custom":
        base_url = settings.base_url.strip()
        if not base_url:
            raise ValueError("Provide base_url for the custom provider")
        key = settings.api_key.strip() or "token-placeholder"
        return LLMClient(api_key=key, model=settings.model, base_url=base_url, extra_headers=settings.headers)
    return LLMClient(api_key=settings.api_key.strip(), model=settings.model, base_url=settings.base_url or "", extra_headers=settings.headers)

# ---------- Similarity ----------
def max_similarities(texts: List[str], names: List[str]) -> Tuple[List[float], List[str], pd.DataFrame]:
    n = len(texts)
    norm = [normalize_for_sim(t) for t in texts]
    max_sim = [0.0] * n
    near_name = [""] * n
    pairs = []
    for i in range(n):
        for j in range(i+1, n):
            s = float(fuzz.token_set_ratio(norm[i], norm[j]))
            pairs.append((names[i], names[j], s))
            if s > max_sim[i]:
                max_sim[i] = s; near_name[i] = names[j]
            if s > max_sim[j]:
                max_sim[j] = s; near_name[j] = names[i]
    pairs_df = pd.DataFrame(pairs, columns=["FileA","FileB","Similarity"]).sort_values("Similarity", ascending=False)
    return max_sim, near_name, pairs_df

# ---------- Screening / Scoring ----------
def compute_scores_table(base_rows: List[Dict], criteria: List[Criterion], dup_threshold: int) -> pd.DataFrame:
    df = pd.DataFrame(base_rows)

    crit_cols = [f"{c.name} (0-5)" for c in criteria if f"{c.name} (0-5)" in df.columns]
    if crit_cols:
        df["Coverage"] = (df[crit_cols] > 0.0).sum(axis=1) / max(1, len(crit_cols))
        pct = df[crit_cols].rank(pct=True)
        pct.columns = [c.replace(" (0-5)", "::Pct") for c in pct.columns]
        df = pd.concat([df, pct], axis=1)
    else:
        df["Coverage"] = 0.0

    weights = {c.name: float(c.weight) for c in criteria}
    sumw = sum(weights.values()) or 1.0

    w_sum = np.zeros(len(df))
    for c in criteria:
        pcol = f"{c.name}::Pct"
        if pcol in df:
            w_sum += weights[c.name] * df[pcol].fillna(0).to_numpy()

    df["CompositeScore"] = 100.0 * (0.75 * (w_sum / sumw) + 0.25 * df["Coverage"])

    def bucket(x): return "A" if x>=80 else ("B" if x>=60 else "C")
    df["PriorityBucket"] = df["CompositeScore"].apply(bucket)

    def clamp_local(s: str, n: int) -> str:
        s = re.sub(r"\s+", " ", (s or "").strip())
        return (s[:n-1] + "…") if len(s) > n else s

    def calc_comment(row):
        fio = clamp_local(row.get("FullName", ""), 80)
        spec = clamp_local(row.get("Specialization", ""), 100)
        cov = row.get("Coverage", 0.0)

        pct_cols = [(k.replace("::Pct",""), k) for k in df.columns if k.endswith("::Pct")]
        top = sorted([(crit, float(row[pcol])) for crit, pcol in pct_cols], key=lambda x: x[1], reverse=True)[:3]

        reason = row.get("_Reasoning", {}) or {}
        strengths_parts, examples_parts = [], []
        for crit, _ in top:
            score_val = row.get(f"{crit} (0-5)", 0.0)
            rtxt = clamp_local(str(reason.get(crit, "")), 280)
            if rtxt:
                examples_parts.append(f"{crit} ({score_val:.1f}): {rtxt}")
            strengths_parts.append(f"{crit} ({score_val:.1f})")

        gaps = []
        for c in [c.name for c in criteria]:
            sc = float(row.get(f"{c} (0-5)", 0.0))
            if sc <= 1.5:
                gap_reason = clamp_local(str(reason.get(c, "")), 160)
                gaps.append(f"{c} ({sc:.1f})" + (f": {gap_reason}" if gap_reason else ""))

        risks = []
        if float(row.get("SimilarityMax", 0)) >= dup_threshold: risks.append("possible duplicate by similarity")
        if not row.get("Email"): risks.append("missing email")
        if not row.get("Phone"): risks.append("missing phone")

        strengths_txt = ", ".join(strengths_parts) if strengths_parts else "no standout strengths"
        examples_txt = " | ".join(examples_parts) if examples_parts else ""
        gaps_txt = "; ".join(gaps) if gaps else "—"

        return (
            f"{('Candidate: ' + fio + '. ') if fio else ''}"
            f"{('Specialization: ' + spec + '. ') if spec else ''}"
            f"Total {row['CompositeScore']:.0f}/100. Coverage {cov:.0%}. "
            f"Strengths: {strengths_txt}. "
            f"{('Examples: ' + examples_txt + '. ') if examples_txt else ''}"
            f"Gaps: {gaps_txt}."
            f"{(' Risks: ' + ', '.join(risks) + '.') if risks else ''}"
        )

    df["CalcComment"] = df.apply(calc_comment, axis=1)
    return df

# ---------- Checkpoint ----------
def load_checkpoint(cp_path: str) -> Dict[str, Dict]:
    if not cp_path or not os.path.exists(cp_path): return {}
    cache: Dict[str, Dict] = {}
    with open(cp_path, "r", encoding="utf-8") as f:
        for line in f:
            line=line.strip()
            if not line: continue
            try:
                obj = json.loads(line)
                cache[obj["hash"]] = obj["data"]
            except Exception:
                continue
    return cache

def append_checkpoint(cp_path: str, file_hash: str, data: Dict):
    with open(cp_path, "a", encoding="utf-8") as f:
        f.write(json.dumps({"hash": file_hash, "data": data}, ensure_ascii=False) + "\n")

def reset_checkpoint(cp_path: str):
    if cp_path and os.path.exists(cp_path): os.remove(cp_path)


# ---------- Validation & intake helpers ----------
def validate_inputs(settings: LLMSettings, role_ctx: RoleContext, intake_mode: str, files, link_inputs: List[str]):
    """Stop execution with clear UI errors when required inputs are missing."""
    if settings.provider == "openai" and not settings.api_key:
        st.error("Provide an OpenAI API Key")
        st.stop()
    if settings.provider == "openrouter" and not settings.api_key:
        st.error("Provide an OpenRouter API Key")
        st.stop()
    if settings.provider == "custom" and not settings.base_url.strip():
        st.error("Custom providers require a base_url")
        st.stop()
    if not role_ctx.description.strip():
        st.error("Role description is required: add a short vacancy/context paragraph")
        st.stop()
    if not role_ctx.criteria:
        st.error("Please provide valid criteria")
        st.stop()
    if intake_mode == "Select an option":
        st.error("Choose how to provide resumes: upload files or paste cloud links")
        st.stop()
    if intake_mode == "Upload files" and not files:
        st.error("Upload at least one resume file")
        st.stop()
    if intake_mode == "Cloud links" and not link_inputs:
        st.error("Add at least one direct link to resume files or pages")
        st.stop()


def collect_incoming_items(intake_mode: str, files, link_inputs: List[str]) -> List[Dict[str, object]]:
    """Normalize uploaded files or remote links into a single list of work items."""
    incoming_items: List[Dict[str, object]] = []
    if intake_mode == "Upload files":
        for f in (files or []):
            try:
                b = f.getvalue()
                incoming_items.append({"kind": "file", "name": f.name, "bytes": b})
            except Exception as e:
                st.error(f"{f.name}: failed to read — {e}")
    elif intake_mode == "Cloud links":
        if link_inputs:
            st.markdown("#### Downloading resumes from links")
            dl_bar = st.progress(0.0)
            for i, u in enumerate(link_inputs, start=1):
                try:
                    fname, data, ct = stream_download(u)
                    incoming_items.append({"kind": "url", "name": fname, "bytes": data, "url": u, "content_type": ct})
                except Exception as e:
                    st.warning(f"Could not download {u}: {e}")
                finally:
                    dl_bar.progress(i / len(link_inputs))
    return incoming_items


def parse_and_dedupe_items(incoming_items: List[Dict[str, object]], status_holder, progress_bar) -> Tuple[List[ResumeArtifact], set, set]:
    """Read bytes into text, convert HTML when needed, and drop duplicates early."""
    parsed_items: List[ResumeArtifact] = []
    seen_file_hash, seen_text_hash = set(), set()
    total_items = len(incoming_items)

    for i, item in enumerate(incoming_items, start=1):
        disp_name = item.get("name") or item.get("url") or f"item_{i}"
        status_holder.text(f"Reading source: {disp_name} ({i}/{total_items})")
        b = item["bytes"]
        fh = sha1_bytes(b)
        if fh in seen_file_hash:
            progress_bar.progress(i / total_items)
            continue

        content_type = (item.get("content_type") or "").lower()
        ext = os.path.splitext(disp_name)[1].lower()

        try:
            if content_type.startswith("text/html") or ext in {".html", ".htm"}:
                text = html_to_text(b, base_url=item.get("url", ""))
            else:
                if ext not in ALLOWED_EXT and ext == "":
                    disp_name = disp_name + ".pdf"
                    ext = ".pdf"
                text = read_file_text(disp_name, b)
        except Exception as e:
            st.error(f"{disp_name}: failed to read — {e}")
            progress_bar.progress(i / total_items)
            continue

        th = sha1_text(normalize_for_sim(text))
        if th in seen_text_hash:
            progress_bar.progress(i / total_items)
            continue

        parsed_items.append(
            ResumeArtifact(
                id=fh,
                file_hash=fh,
                text_hash=th,
                name=disp_name,
                text=text,
                url=item.get("url", ""),
            )
        )

        seen_file_hash.add(fh)
        seen_text_hash.add(th)
        progress_bar.progress(i / total_items)

    return parsed_items, seen_file_hash, seen_text_hash

# ---------- UI ----------
st.set_page_config(page_title="🍵 Matchya — Hire faster", layout="wide")
st.markdown(
    """
    <style>
    section[data-testid='stSidebar'] {width: 27rem;}
    section[data-testid='stSidebar'] > div:first-child {width: 27rem;}
    </style>
    """,
    unsafe_allow_html=True,
)
st.title("🍵 Matchya")
st.caption("Hire faster")

with st.sidebar:
    st.header("⚙️ LLM")
    provider = st.selectbox(
        "LLM provider",
        ["openai", "openrouter", "lmstudio", "custom"],
        format_func=lambda x: {
            "openai": "OpenAI (cloud)",
            "openrouter": "OpenRouter (cloud)",
            "lmstudio": "LM Studio (local)",
            "custom": "Custom base_url",
        }[x],
    )
    api_key = st.text_input(
        "API Key",
        type="password",
        help="Leave blank for LM Studio if you rely on the default 'lm-studio' token.",
    )
    default_lm_url = "http://localhost:1234/v1"
    custom_base_url = ""
    if provider in {"lmstudio", "custom"}:
        custom_base_url = st.text_input(
            "Base URL (LM Studio/custom)",
            value=default_lm_url if provider == "lmstudio" else "",
            help="OpenAI-compatible endpoint. LM Studio default port is 1234.",
        )
    model_options = fetch_provider_models(provider, api_key, custom_base_url)
    model_name = st.selectbox("Model", model_options, index=0)

    llm_headers = {"HTTP-Referer": "https://github.com/user/portfolio", "X-Title": "Matchya"} if provider == "openrouter" else None
    llm_settings = LLMSettings(
        provider=provider,
        api_key=api_key,
        model=model_name,
        base_url=custom_base_url,
        headers=llm_headers,
    )

    st.header("📌 Role & criteria")
    role_desc = st.text_area("Role / vacancy description (required)", height=140)

    st.subheader("Key skills / criteria (required)")
    default_criteria = [
        {"name": "Domain experience", "weight": 2.0, "keywords": ["experience", "domain"]},
        {"name": "Hard skills", "weight": 1.8, "keywords": ["stack", "tech"]},
        {"name": "Soft skills", "weight": 1.2, "keywords": ["communication", "teamwork"]},
        {"name": "Achievements", "weight": 1.4, "keywords": ["results", "impact"]},
    ]
    crit_state_key = "criteria_table"
    if crit_state_key not in st.session_state:
        st.session_state[crit_state_key] = [
            {
                "Criterion": c["name"],
                "Weight": c["weight"],
                "Keywords": ", ".join(c.get("keywords") or []),
            }
            for c in default_criteria
        ]

    def _generate_criteria():
        if not role_desc.strip():
            st.error("Add a role description first to generate skills automatically.")
            return
        try:
            generator_client = create_llm_client(llm_settings)
            generated = generator_client.suggest_criteria(role_desc)
            if not generated:
                st.warning("LLM returned no skills. Please fill manually.")
                return
            st.session_state[crit_state_key] = [
                {"Criterion": c.name, "Weight": c.weight, "Keywords": ", ".join(c.keywords)}
                for c in generated
            ]
            st.session_state["criteria_generated_ok"] = True
        except Exception as e:
            st.error(f"Could not generate skills: {e}")

    crit_cols = st.columns([1, 3])
    with crit_cols[0]:
        if st.button("⚡️ Generate skills", use_container_width=True):
            _generate_criteria()
        st.caption("Optional: draft criteria from the description, then tweak below.")

    editor_value = st.data_editor(
        st.session_state[crit_state_key],
        column_config={
            "Criterion": st.column_config.TextColumn("Criterion", help="One skill per row"),
            "Weight": st.column_config.NumberColumn(
                "Weight (0.0–3.0)", min_value=0.0, step=0.1, format="%.1f"
            ),
            "Keywords": st.column_config.TextColumn(
                "Keywords (comma-separated)",
                help="Optional hints to keep the LLM focused",
                width="medium",
            ),
        },
        num_rows="dynamic",
        hide_index=True,
        use_container_width=True,
        key="criteria_editor",
    )

    st.session_state[crit_state_key] = (
        editor_value.to_dict("records") if hasattr(editor_value, "to_dict") else editor_value
    )

    if st.session_state.pop("criteria_generated_ok", False):
        st.success("Skills generated and inserted below")

    criteria: List[Criterion] = []
    try:
        crit_records = st.session_state.get(crit_state_key, [])
        parsed_records: List[Dict[str, object]] = (
            crit_records.to_dict("records") if hasattr(crit_records, "to_dict") else list(crit_records)
        )
        for row in parsed_records:
            name = str(row.get("Criterion", "")).strip()
            if not name:
                continue
            weight_val = float(row.get("Weight") or 0.0)
            kw_raw = row.get("Keywords") or ""
            if isinstance(kw_raw, list):
                keywords = [str(k).strip() for k in kw_raw if str(k).strip()]
            else:
                keywords = [k.strip() for k in str(kw_raw).split(",") if k.strip()]
            criteria.append(Criterion(name=name, weight=weight_val, keywords=keywords))
    except Exception as e:
        st.error(f"Criteria parsing error: {e}")

    role_ctx = RoleContext(description=role_desc, criteria=criteria)

    st.subheader("Duplicates")
    dup_threshold = st.slider("Similarity threshold (duplicate risk)", min_value=70, max_value=100, value=90, step=1)

    st.subheader("Output")
    save_path = st.text_input("Save XLSX path (server)", value="resume_ranking.xlsx")
    add_pairs = st.checkbox("Include SimilarityPairs sheet (top 200)", value=True)

    st.subheader("Checkpoint")
    cp_path = st.text_input("Checkpoint file (.jsonl)", value="resume_ranker_checkpoint.jsonl")
    colA, colB = st.columns(2)
    with colA:
        resume_from_cp = st.checkbox("Resume from checkpoint", value=True)
    with colB:
        if st.button("♻️ Reset checkpoint"):
            reset_checkpoint(cp_path); st.success("Checkpoint removed.")

st.markdown("## 📥 Resume intake")
intake_mode = st.selectbox("Choose how to provide resumes", ["Select an option", "Upload files", "Cloud links"])
files: List = []
link_inputs: List[str] = []

if intake_mode == "Upload files":
    files = st.file_uploader(
        "Files (PDF/DOCX/TXT/MD/RTF)",
        type=[ext[1:] for ext in ALLOWED_EXT],
        accept_multiple_files=True,
    )
elif intake_mode == "Cloud links":
    urls_raw = st.text_area(
        "One link per line (direct links to resume files or HTML pages)",
        placeholder="https://...",
    )
    link_inputs = [normalize_url(u) for u in urls_raw.splitlines() if normalize_url(u)]

run = st.button("🚀 Process and export XLSX")

# ---------- Main ----------
if run:
    validate_inputs(llm_settings, role_ctx, intake_mode, files, link_inputs)

    try:
        client = create_llm_client(llm_settings)
    except Exception as e:
        st.error(f"LLM client could not be created: {e}")
        st.stop()
    cache = load_checkpoint(cp_path) if resume_from_cp else {}

    rows: List[Dict] = []
    texts: List[str] = []
    filenames: List[str] = []
    criteria = role_ctx.criteria

    incoming_items = collect_incoming_items(intake_mode, files, link_inputs)

    status = st.empty()
    st.markdown("#### Progress")
    progress_bar = st.progress(0.0)

    parsed_items, _, _ = parse_and_dedupe_items(incoming_items, status, progress_bar)

    kept_after_parse = len(parsed_items)
    status.text(f"Parsing finished. Sources: {len(incoming_items)}. After local dedup: {kept_after_parse}.")

    if not parsed_items:
        st.warning("No usable resumes after parsing/dedup"); st.stop()

    progress_bar.progress(0.0)

    total_llm = len(parsed_items)
    for idx, it in enumerate(parsed_items, start=1):
        status.text(f"LLM {idx}/{total_llm}: {it.name}")

        if "pack" not in cache.get(it.file_hash, {}):
            try:
                pack = client.score_and_extract_single(
                    resume={"id": it.id, "text": it.text},
                    role_desc=role_ctx.description,
                    criteria=criteria,
                )
            except Exception as e:
                st.error(f"LLM error on {it.name}: {e}")
                pack = {}

            obj = cache.get(it.file_hash, {})
            obj["pack"] = {
                "id": it.id,
                "full_name": "",
                "specialization_main": "",
                "specialization_alt": [],
                "emails": [],
                "phones": [],
                "scores": {},
                "reasoning": {},
                **(pack or {}),
            }
            cache[it.file_hash] = obj
            append_checkpoint(cp_path, it.file_hash, obj)

        cobj = cache[it.file_hash]["pack"]
        text = it.text

        emails_all = list(cobj.get("emails") or []) or EMAIL_RE.findall(text)
        phones_all = list(cobj.get("phones") or []) or PHONE_CAND_RE.findall(text)
        email_final = emails_all[0] if emails_all else ""
        phone_final = best_phone(phones_all)

        fio_llm = (cobj.get("full_name") or "").strip()
        fio_val = fio_llm if fio_llm else guess_full_name(text)

        specialization = (cobj.get("specialization_main") or "").strip()
        if not specialization:
            lines = [l.strip() for l in text.splitlines() if l.strip()][:20]
            for l in lines[:15]:
                if "@" in l or re.search(r"https?://", l): continue
                if EMAIL_RE.search(l) or PHONE_CAND_RE.search(l): continue
                if len(l) <= 140 and (l.istitle() or re.search(r"[A-Za-zА-Яа-я ]{6,}", l)):
                    specialization = l; break

        score_map = {k: float(cobj.get("scores", {}).get(k, 0.0)) for k in [c.name for c in criteria]}
        base = {
            "File": it.name,
            "FullName": fio_val,
            "Specialization": specialization,
            "Email": email_final,
            "Phone": phone_final,
            "FullText": text,
            "SourceURL": it.url,
            "_FileHash": it.file_hash, "_TextHash": it.text_hash,
            "_Reasoning": cobj.get("reasoning", {}),
        }
        for c in criteria:
            base[f"{c.name} (0-5)"] = score_map.get(c.name, 0.0)

        rows.append(base)
        texts.append(text); filenames.append(it.name)

        progress_bar.progress(idx/total_llm)

    if not rows:
        st.warning("No successful LLM results"); st.stop()

    # -------- Pass 3: similarity + email/phone duplicate removal --------
    sim_max, sim_near, pairs_df = max_similarities(texts, filenames)
    for idx, row in enumerate(rows):
        row["SimilarityMax"] = round(sim_max[idx], 1)
        row["NearDuplicateOf"] = sim_near[idx]

    by_email: Dict[str,int] = {}
    by_phone: Dict[str,int] = {}
    keep_mask = [True]*len(rows)
    def norm_email(e: str) -> str: return (e or "").strip().lower()
    def norm_phone(p: str) -> str: return "".join(d for d in (p or "") if d.isdigit())

    for i,r in enumerate(rows):
        e = norm_email(r.get("Email",""))
        if e:
            if e in by_email: keep_mask[i] = False
            else: by_email[e] = i
    for i,r in enumerate(rows):
        if not keep_mask[i]: continue
        p = norm_phone(r.get("Phone",""))
        if p:
            if p in by_phone: keep_mask[i] = False
            else: by_phone[p] = i

    if not pairs_df.empty:
        name_to_idx = {rows[i]["File"]: i for i in range(len(rows))}
        for _, r in pairs_df.iterrows():
            a, b, s = r["FileA"], r["FileB"], float(r["Similarity"])
            ia, ib = name_to_idx.get(a), name_to_idx.get(b)
            if ia is None or ib is None: continue
            if not keep_mask[ia] or not keep_mask[ib]: continue
            if s >= 100.0 or s >= dup_threshold:
                drop = ib if ia < ib else ia
                keep_mask[drop] = False

    rows = [r for i,r in enumerate(rows) if keep_mask[i]]

    # -------- Tables and export --------
    df = compute_scores_table(rows, criteria, dup_threshold)

    show_cols = [
        "File","FullName","Specialization","Email","Phone",
        "SourceURL",
        "Coverage","CompositeScore","PriorityBucket","SimilarityMax","NearDuplicateOf","CalcComment"
    ]
    crit_cols = [f"{c.name} (0-5)" for c in criteria if f"{c.name} (0-5)" in df.columns]
    final_df = df[[c for c in show_cols if c in df.columns] + crit_cols].copy()
    final_df.insert(0, "Rank", range(1, len(final_df)+1))
    final_df = final_df.sort_values(["CompositeScore","SimilarityMax"], ascending=[False, False]).reset_index(drop=True)
    final_df["Rank"] = range(1, len(final_df)+1)

    # Stats sheets
    stats = []
    for c in criteria:
        pcol = f"{c.name}::Pct"
        if pcol in df:
            s = df[pcol].dropna()
            stats.append({
                "Criterion": c.name, "Weight": float(c.weight),
                "MedianPct": float(s.median()) if len(s) else 0.0,
                "P10": float(s.quantile(0.10)) if len(s) else 0.0,
                "P90": float(s.quantile(0.90)) if len(s) else 0.0,
                "CoverageShare": float((df[f"{c.name} (0-5)"]>0).mean()) if f"{c.name} (0-5)" in df else 0.0
            })
    crit_stats_df = pd.DataFrame(stats)
    similarity_pairs_df = pairs_df.head(200).copy() if add_pairs and not pairs_df.empty else pd.DataFrame(columns=["FileA","FileB","Similarity"])

    logic_text = (
        "How scores are produced:\n"
        "• LLM scores 0–5 per criterion with reasoning — one resume per request to avoid cross-talk.\n"
        "• Full name comes from the LLM with a light text heuristic fallback.\n"
        "• Each request includes the role description and the criteria table (sent as JSON) — scores rely on this context.\n"
        "• Composite = 0.75×weighted percentiles + 0.25×coverage.\n"
        "• Duplicates: identical files/texts and repeated email/phone are dropped; Similarity ≥ threshold is flagged as risk.\n"
        "• Comment includes name/specialization, top strengths with excerpts, gaps (low scores), and risks."
    )
    config_df = pd.DataFrame({
        "Key":[
            "Model","RoleDescription","DuplicateThreshold","CheckpointFile",
            "TotalUploaded","KeptAfterLocalDedup","BatchSize","NumBatches","HumanLogic"
        ],
        "Value":[
            model_name,
            role_desc.strip()[:160] + ("…" if len(role_desc.strip()) > 160 else ""),
            str(dup_threshold), cp_path,
            str(total_items), str(len(parsed_items)), "1 (no batching)", str(len(parsed_items)), logic_text
        ]
    })

    # Write XLSX with formatting
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.utils import get_column_letter

    DATA_BAR_COLOR = Color("FF63BE7B")

    xlsx_buf = io.BytesIO()
    with pd.ExcelWriter(xlsx_buf, engine="openpyxl") as xw:
        final_df.to_excel(xw, sheet_name="Ranking", index=False)
        crit_stats_df.to_excel(xw, sheet_name="CriteriaStats", index=False)
        similarity_pairs_df.to_excel(xw, sheet_name="SimilarityPairs", index=False)
        config_df.to_excel(xw, sheet_name="Config", index=False)

        wb = xw.book
        ws = wb["Ranking"]

        thin = Side(border_style="thin", color="DDDDDD")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.border = border_all
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in r:
                cell.border = border_all
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        stripe = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
        for row in range(2, ws.max_row+1, 2):
            for col in range(1, ws.max_column+1):
                ws.cell(row=row, column=col).fill = stripe

        for col_idx in range(1, ws.max_column+1):
            col = get_column_letter(col_idx)
            max_len = max(len(str(ws[f"{col}{r}"].value)) if ws[f"{col}{r}"].value is not None else 0 for r in range(1, ws.max_row+1))
            ws.column_dimensions[col].width = min(50, max(12, max_len + 2))

        headers = {cell.value: cell.col_idx for cell in ws[1]}
        def col_letter(name: str) -> str: return get_column_letter(headers[name])

        if "CompositeScore" in headers:
            c = col_letter("CompositeScore")
            ws.conditional_formatting.add(
                f"{c}2:{c}{ws.max_row}",
                ColorScaleRule(start_type="min", start_color="F8696B",
                               mid_type="percentile", mid_value=50, mid_color="FFEB84",
                               end_type="max", end_color="63BE7B")
            )

        for h, idx in headers.items():
            if h.endswith(" (0-5)"):
                col = get_column_letter(idx)
                rule = DataBarRule(
                    start_type="num", start_value=0,
                    end_type="num", end_value=5,
                    color=DATA_BAR_COLOR,
                    showValue=True
                )
                ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", rule)

        if "PriorityBucket" in headers:
            bc = col_letter("PriorityBucket")
            for row in range(2, ws.max_row+1):
                v = str(ws[f"{bc}{row}"].value or "")
                fill = None
                if v == "A": fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
                elif v == "B": fill = PatternFill(start_color="FFF5CC", end_color="FFF5CC", fill_type="solid")
                elif v == "C": fill = PatternFill(start_color="FDE7E9", end_color="FDE7E9", fill_type="solid")
                if fill:
                    for col in range(1, ws.max_column+1):
                        ws.cell(row=row, column=col).fill = fill

        if "SimilarityMax" in headers:
            sc = col_letter("SimilarityMax")
            for row in range(2, ws.max_row+1):
                try:
                    val = float(ws[f"{sc}{row}"].value)
                    if val >= dup_threshold:
                        fill = PatternFill(start_color="FFE8AA", end_color="FFE8AA", fill_type="solid")
                        for col in range(1, ws.max_column+1):
                            ws.cell(row=row, column=col).fill = fill
                        ws[f"{sc}{row}"].font = Font(bold=True)
                except Exception:
                    pass

    data = xlsx_buf.getvalue()

    # Save to server
    try:
        if save_path:
            with open(save_path, "wb") as f: f.write(data)
            st.success(f"Saved on server: {save_path}")
    except Exception as e:
        st.warning(f"Could not save on server: {e}")

    # Download
    st.download_button("⬇️ Download XLSX", data=data,
                       file_name=os.path.basename(save_path) or "resume_ranking.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Footer branding
st.markdown("---")
st.caption("© 🍵 Matchya — Hire faster")

