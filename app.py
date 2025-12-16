"""
🗑️ clean_the_garbage.exe — Resume Ranker (Ultimate v5+)
Company: Lola, Liza & Partners LLC

Функции:
- LLM (OpenAI, по умолчанию gpt-4o-mini) извлекает ФИО, специализацию (+альтернативы), контакты,
  оценивает по критериям 0–5 и даёт краткие пояснения — ВСЁ за один батч-запрос (до 5 резюме).
- Устойчивый композитный скоринг: 0.75*wP(перцентили) + 0.25*Coverage.
- Жёсткая дедупликация: одинаковые файлы/тексты, одинаковые email/телефоны, Similarity >= порога (100% — всегда дубликат).
- Выгрузка ТОЛЬКО в XLSX. Бордеры, жирные заголовки, шкалы, подсветка строк по бакетам/рискам.
- Чекпоинт JSONL по sha1 от байтов — безопасное возобновление.

Установка:
    pip install streamlit pdfminer.six python-docx rapidfuzz pandas openpyxl pydantic tenacity openai

Запуск:
    streamlit run app.py
"""

from __future__ import annotations
import io, os, re, json, hashlib
from typing import List, Dict, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from pydantic import BaseModel, Field
from rapidfuzz import fuzz

# ---------- Parsers ----------
try:
    from pdfminer.high_level import extract_text as pdf_extract_text
except Exception:
    pdf_extract_text = None
try:
    import docx
except Exception:
    docx = None

ALLOWED_EXT = {".pdf", ".docx", ".txt", ".md", ".rtf"}
TOKEN_SPLIT = re.compile(r"[\W_]+", re.UNICODE)
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
PHONE_CAND_RE = re.compile(r"(?:\+?\d[\d\-\s()\./]{6,}\d)")

# ---------- Models ----------
class Criterion(BaseModel):
    name: str
    weight: float = Field(1.0, ge=0)
    keywords: List[str] = Field(default_factory=list)

class LLMScores(BaseModel):
    scores: Dict[str, float]
    reasoning: Dict[str, str] = Field(default_factory=dict)

# ---------- Utils ----------
def read_file_text(filename: str, bytes_data: bytes) -> str:
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".pdf":
        if not pdf_extract_text:
            raise RuntimeError("pdfminer.six не установлен. pip install pdfminer.six")
        with io.BytesIO(bytes_data) as bio:
            return pdf_extract_text(bio)
    if ext == ".docx":
        if not docx:
            raise RuntimeError("python-docx не установлен. pip install python-docx")
        with io.BytesIO(bytes_data) as bio:
            d = docx.Document(bio)
            return "\n".join(p.text for p in d.paragraphs)
    if ext in {".txt", ".md", ".rtf"}:
        try:
            return bytes_data.decode("utf-8", errors="ignore")
        except Exception:
            return bytes_data.decode("latin-1", errors="ignore")
    raise ValueError(f"Неподдерживаемый формат: {ext}")

def sha1_bytes(b: bytes) -> str:
    return hashlib.sha1(b).hexdigest()

def sha1_text(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8", errors="ignore")).hexdigest()

def normalize_for_sim(t: str) -> str:
    t = re.sub(r"\s+", " ", t.lower())
    return re.sub(r"[^a-zа-я0-9 ]+", " ", t)

def normalize_phone_digits(s: str) -> str:
    return "".join(ch for ch in s if ch.isdigit() or ch == "+")

def best_phone(phones: List[str]) -> str:
    cleaned = []
    for p in phones:
        norm = normalize_phone_digits(p)
        digits = "".join(d for d in norm if d.isdigit())
        if 10 <= len(digits) <= 15:
            cleaned.append(norm)
    if not cleaned:
        return ""
    cleaned.sort(key=lambda x: (not x.startswith("+"), -len(x)))
    return cleaned[0]

# --- эвристика для ФИО (фолбэк) ---
FIO_RE = re.compile(r"\b[А-ЯЁ][а-яё]+ [А-ЯЁ][а-яё]+(?: [А-ЯЁ][а-яё]+)?\b")

def guess_fio(text: str) -> str:
    lines = [l.strip() for l in text.splitlines() if l.strip()][:30]
    for l in lines:
        if EMAIL_RE.search(l) or PHONE_CAND_RE.search(l) or "http" in l.lower():
            continue
        m = FIO_RE.search(l)
        if m:
            return m.group(0)
    return ""

def chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i+n]

def clamp(s: str, max_len: int) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip())
    return (s[:max_len-1] + "…") if len(s) > max_len else s

# ---------- LLM ----------
class OpenAIClientWrapper:
    def __init__(self, api_key: str, model: str = "gpt-4o-mini"):
        from openai import OpenAI
        self.client = OpenAI(api_key=api_key)
        self.model = model

    def score_and_extract_batch(
        self,
        resumes: List[Dict[str, str]],
        role_desc: str,
        criteria: List[Criterion],
        job_title: str = ""
    ) -> List[Dict[str, object]]:
        """
        Один вызов на партию до 5 резюме.
        Вход: resumes = [{id: str, text: str}, ...]
        Выход: [
          {
            "id": str,
            "full_name": str,
            "specialization_main": str,
            "specialization_alt": [str],
            "emails": [str],
            "phones": [str],
            "scores": {criterion: float},
            "reasoning": {criterion: str}
          }, ...
        ]
        """
        system = (
            "Ты — ассистент HR. Для КАЖДОГО резюме из списка:\n"
            "1) Извлеки полное имя кандидата (ФИО) как 'full_name' (если нет — пусто).\n"
            "2) Извлеки специализацию: 'specialization_main' (до 80 символов) и до 3 альтернатив.\n"
            "3) Извлеки контакты: 'emails' и 'phones' (желательно в международном формате).\n"
            "4) Оцени по критериям 0..5 и дай краткие, но информативные пояснения по каждому критерию.\n"
            "Возвращай JSON строго по схеме для всех входов."
        )
        payload = {
            "role_title": job_title,
            "role_description": role_desc,
            "criteria": [c.model_dump() for c in criteria],
            "resumes": [{"id": r["id"], "text": r["text"][:18000]} for r in resumes][:5]
        }

        schema = {
            "type": "object",
            "properties": {
                "results": {
                    "type": "array",
                    "maxItems": 5,
                    "items": {
                        "type": "object",
                        "properties": {
                            "id": {"type": "string"},
                            "full_name": {"type": "string"},
                            "specialization_main": {"type": "string"},
                            "specialization_alt": {"type": "array", "items": {"type": "string"}},
                            "emails": {"type": "array", "items": {"type": "string"}},
                            "phones": {"type": "array", "items": {"type": "string"}},
                            "scores": {"type": "object", "additionalProperties": {"type": "number"}},
                            "reasoning": {"type": "object", "additionalProperties": {"type": "string"}}
                        },
                        "required": [
                            "id","full_name","specialization_main","emails","phones","scores"
                        ],
                        "additionalProperties": False
                    }
                }
            },
            "required": ["results"],
            "additionalProperties": False
        }

        resp = self.client.chat.completions.create(
            model=self.model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": json.dumps(payload, ensure_ascii=False)}
            ],
            response_format={"type": "json_schema", "json_schema": {"name": "ResumeBatch", "schema": schema}},
            temperature=0.1,
        )
        content = resp.choices[0].message.content
        data = json.loads(content)
        return data.get("results", [])

# ---------- Similarity ----------
def max_similarities(texts: List[str], names: List[str]) -> Tuple[List[float], List[str], pd.DataFrame]:
    n = len(texts)
    norm = [normalize_for_sim(t) for t in texts]
    max_sim = [0.0] * n
    near_name = [""] * n
    pairs = []
    for i in range(n):
        for j in range(i+1, n):
            s = float(fuzz.token_set_ratio(norm[i], norm[j]))
            pairs.append((names[i], names[j], s))
            if s > max_sim[i]:
                max_sim[i] = s; near_name[i] = names[j]
            if s > max_sim[j]:
                max_sim[j] = s; near_name[j] = names[i]
    pairs_df = pd.DataFrame(pairs, columns=["FileA","FileB","Similarity"]).sort_values("Similarity", ascending=False)
    return max_sim, near_name, pairs_df

# ---------- Screening / Scoring ----------
def compute_scores_table(base_rows: List[Dict], criteria: List[Criterion], dup_threshold: int) -> pd.DataFrame:
    df = pd.DataFrame(base_rows)

    # Coverage и перцентильные колонки
    crit_cols = [f"{c.name} (0-5)" for c in criteria if f"{c.name} (0-5)" in df.columns]
    if crit_cols:
        df["Coverage"] = (df[crit_cols] > 0.0).sum(axis=1) / max(1, len(crit_cols))
        pct = df[crit_cols].rank(pct=True)
        pct.columns = [c.replace(" (0-5)", "::Pct") for c in pct.columns]
        df = pd.concat([df, pct], axis=1)
    else:
        df["Coverage"] = 0.0

    # Веса критериев
    weights = {c.name: float(c.weight) for c in criteria}
    sumw = sum(weights.values()) or 1.0

    # Взвешенная сумма перцентилей
    w_sum = np.zeros(len(df))
    for c in criteria:
        pcol = f"{c.name}::Pct"
        if pcol in df:
            w_sum += weights[c.name] * df[pcol].fillna(0).to_numpy()

    # Итог: 0.75 перцентили + 0.25 покрытие (без RecencyBoost)
    df["CompositeScore"] = 100.0 * (0.75 * (w_sum / sumw) + 0.25 * df["Coverage"])

    def bucket(x): return "A" if x>=80 else ("B" if x>=60 else "C")
    df["PriorityBucket"] = df["CompositeScore"].apply(bucket)

    # Развёрнутый комментарий
    def clamp_local(s: str, n: int) -> str:
        s = re.sub(r"\s+", " ", (s or "").strip())
        return (s[:n-1] + "…") if len(s) > n else s

    def calc_comment(row):
        fio = clamp_local(row.get("ФИО", ""), 80)
        spec = clamp_local(row.get("Специализация", ""), 100)
        cov = row.get("Coverage", 0.0)

        pct_cols = [(k.replace("::Pct",""), k) for k in df.columns if k.endswith("::Pct")]
        top = sorted([(crit, float(row[pcol])) for crit, pcol in pct_cols], key=lambda x: x[1], reverse=True)[:3]

        reason = row.get("_Reasoning", {}) or {}
        strengths_parts, examples_parts = [], []
        for crit, _ in top:
            score_val = row.get(f"{crit} (0-5)", 0.0)
            rtxt = clamp_local(str(reason.get(crit, "")), 280)
            if rtxt:
                examples_parts.append(f"{crit} ({score_val:.1f}): {rtxt}")
            strengths_parts.append(f"{crit} ({score_val:.1f})")

        gaps = []
        for c in [c.name for c in criteria]:
            sc = float(row.get(f"{c} (0-5)", 0.0))
            if sc <= 1.5:
                gap_reason = clamp_local(str(reason.get(c, "")), 160)
                gaps.append(f"{c} ({sc:.1f})" + (f": {gap_reason}" if gap_reason else ""))

        risks = []
        if float(row.get("SimilarityMax", 0)) >= dup_threshold: risks.append("риск дубликата по Similarity")
        if not row.get("Email"): risks.append("нет email")
        if not row.get("Phone"): risks.append("нет телефона")

        strengths_txt = ", ".join(strengths_parts) if strengths_parts else "нет явных сильных сторон"
        examples_txt = " | ".join(examples_parts) if examples_parts else ""
        gaps_txt = "; ".join(gaps) if gaps else "—"

        return (
            f"{('Кандидат: ' + fio + '. ') if fio else ''}"
            f"{('Специализация: ' + spec + '. ') if spec else ''}"
            f"Итог {row['CompositeScore']:.0f}/100. Покрытие {cov:.0%}. "
            f"Сильные стороны: {strengths_txt}. "
            f"{('Примеры: ' + examples_txt + '. ') if examples_txt else ''}"
            f"Пробелы: {gaps_txt}."
            f"{(' Риски: ' + ', '.join(risks) + '.') if risks else ''}"
        )

    df["CalcComment"] = df.apply(calc_comment, axis=1)
    return df

# ---------- Checkpoint ----------
def load_checkpoint(cp_path: str) -> Dict[str, Dict]:
    if not cp_path or not os.path.exists(cp_path): return {}
    cache: Dict[str, Dict] = {}
    with open(cp_path, "r", encoding="utf-8") as f:
        for line in f:
            line=line.strip()
            if not line: continue
            try:
                obj = json.loads(line)
                cache[obj["hash"]] = obj["data"]
            except Exception:
                continue
    return cache

def append_checkpoint(cp_path: str, file_hash: str, data: Dict):
    with open(cp_path, "a", encoding="utf-8") as f:
        f.write(json.dumps({"hash": file_hash, "data": data}, ensure_ascii=False) + "\n")

def reset_checkpoint(cp_path: str):
    if cp_path and os.path.exists(cp_path): os.remove(cp_path)

# ---------- UI ----------
st.set_page_config(page_title="🗑️ clean_the_garbage.exe — Lola, Liza & Partners LLC", layout="wide")
st.title("🗑️ clean_the_garbage.exe")
st.caption("Lola, Liza & Partners LLC — serious screening for massive resume batches")

with st.sidebar:
    st.header("⚙️ LLM")
    api_key = st.text_input("OpenAI API Key", type="password")
    model_name = st.selectbox("Модель", ["gpt-4o-mini","gpt-4o","gpt-4.1-mini"], index=0)

    st.header("📌 Роль/критерии")
    job_title = st.text_input("Название роли (опц.)", value="")
    role_desc = st.text_area("Описание роли/вакансии", height=120)

    st.subheader("Критерии (JSON)")
    default_criteria = [
        {"name": "Релевантность опыту", "weight": 2.0, "keywords": []},
        {"name": "Достижения/результаты", "weight": 1.6, "keywords": []},
        {"name": "Навыки по роли", "weight": 1.8, "keywords": []},
        {"name": "Коммуникация/переговоры", "weight": 1.2, "keywords": []},
        {"name": "Образование/сертификаты", "weight": 0.8, "keywords": []},
    ]
    crit_json = st.text_area("Список критериев", value=json.dumps(default_criteria, ensure_ascii=False, indent=2), height=220)
    criteria: List[Criterion] = []
    try:
        criteria = [Criterion(**c) for c in json.loads(crit_json)]
    except Exception as e:
        st.error(f"Ошибка критериев: {e}")

    st.subheader("Дубликаты")
    dup_threshold = st.slider("Порог похожести (риск дубликата)", min_value=70, max_value=100, value=90, step=1)

    st.subheader("Вывод")
    save_path = st.text_input("Путь сохранения XLSX (на сервере)", value="resume_ranking.xlsx")
    add_pairs = st.checkbox("Лист SimilarityPairs (топ-200)", value=True)

    st.subheader("Чекпоинт")
    cp_path = st.text_input("Файл чекпоинта (.jsonl)", value="resume_ranker_checkpoint.jsonl")
    colA, colB = st.columns(2)
    with colA:
        resume_from_cp = st.checkbox("Возобновлять из чекпоинта", value=True)
    with colB:
        if st.button("♻️ Сбросить чекпоинт"):
            reset_checkpoint(cp_path); st.success("Чекпоинт удалён.")

st.markdown("## 📥 Загрузка резюме")
files = st.file_uploader("Файлы (PDF/DOCX/TXT/MD/RTF)", type=[ext[1:] for ext in ALLOWED_EXT], accept_multiple_files=True)
run = st.button("🚀 Обработать и выгрузить XLSX")

# ---------- Main ----------
if run:
    if not api_key: st.error("Укажите OpenAI API Key"); st.stop()
    if not criteria: st.error("Задайте валидные критерии"); st.stop()
    if not files: st.error("Загрузите файлы"); st.stop()

    client = OpenAIClientWrapper(api_key=api_key, model=model_name)
    cache = load_checkpoint(cp_path) if resume_from_cp else {}

    rows: List[Dict] = []
    texts: List[str] = []
    filenames: List[str] = []
    seen_file_hash, seen_text_hash = set(), set()

    status = st.empty()
    st.markdown("#### Прогресс")
    progress_bar = st.progress(0.0)  # единый прогресс-бар

    # -------- Pass 1: локальный парсинг/дедуп --------
    parsed_items = []  # [{id, fh, th, name, text}]
    total_files = len(files)
    for i, f in enumerate(files, start=1):
        status.text(f"Чтение файла: {f.name} ({i}/{total_files})")
        b = f.getvalue()
        fh = sha1_bytes(b)
        if fh in seen_file_hash:
            progress_bar.progress(i/total_files)
            continue

        try:
            text = read_file_text(f.name, b)
        except Exception as e:
            st.error(f"{f.name}: ошибка чтения — {e}")
            progress_bar.progress(i/total_files)
            continue

        th = sha1_text(normalize_for_sim(text))
        if th in seen_text_hash:
            progress_bar.progress(i/total_files)
            continue

        parsed_items.append({"id": fh, "fh": fh, "th": th, "name": f.name, "text": text})
        seen_file_hash.add(fh); seen_text_hash.add(th)
        progress_bar.progress(i/total_files)

    kept_after_parse = len(parsed_items)
    status.text(f"Парсинг завершён. Файлов: {total_files}. После локальной дедупликации: {kept_after_parse}.")

    if not parsed_items:
        st.warning("Нет успешных результатов после парсинга/дедупликации"); st.stop()

    # -------- Сброс прогресса перед батчами --------
    progress_bar.progress(0.0)

    # -------- Pass 2: LLM батчи (до 5 резюме на запрос) --------
    batches = list(chunks(parsed_items, 5))
    num_batches = len(batches)
    for bi, batch in enumerate(batches, start=1):
        status.text(f"LLM батч {bi}/{num_batches}: {batch[0]['name']} (+{len(batch)-1} ещё)")
        # Проверка чекпоинта
        need_call = False
        batch_payload = []
        for it in batch:
            if "pack" not in cache.get(it["fh"], {}):
                need_call = True
                batch_payload.append({"id": it["fh"], "text": it["text"]})

        if need_call and batch_payload:
            try:
                packs = client.score_and_extract_batch(
                    resumes=batch_payload,
                    role_desc=role_desc or job_title,
                    criteria=criteria,
                    job_title=job_title
                )
            except Exception as e:
                st.error(f"Ошибка LLM на батче {bi}/{num_batches}: {e}")
                packs = []

            by_id = {p["id"]: p for p in packs}
            for it in batch:
                obj = cache.get(it["fh"], {})
                if it["fh"] in by_id:
                    obj["pack"] = by_id[it["fh"]]
                else:
                    obj["pack"] = {
                        "id": it["fh"],
                        "full_name": "",
                        "specialization_main": "",
                        "specialization_alt": [],
                        "emails": [],
                        "phones": [],
                        "scores": {},
                        "reasoning": {}
                    }
                cache[it["fh"]] = obj
                append_checkpoint(cp_path, it["fh"], obj)

        # Сбор строк для вывода (с фолбэками)
        for it in batch:
            cobj = cache[it["fh"]]["pack"]
            text = it["text"]

            emails_all = list(cobj.get("emails") or []) or EMAIL_RE.findall(text)
            phones_all = list(cobj.get("phones") or []) or PHONE_CAND_RE.findall(text)
            email_final = emails_all[0] if emails_all else ""
            phone_final = best_phone(phones_all)

            fio = (cobj.get("full_name") or "").strip() or guess_fio(text)

            specialization = (cobj.get("specialization_main") or "").strip()
            if not specialization:
                lines = [l.strip() for l in text.splitlines() if l.strip()][:20]
                for l in lines[:15]:
                    if "@" in l or re.search(r"https?://", l): continue
                    if EMAIL_RE.search(l) or PHONE_CAND_RE.search(l): continue
                    if len(l) <= 140 and (l.istitle() or re.search(r"[A-Za-zА-Яа-я ]{6,}", l)):
                        specialization = l; break

            score_map = {k: float(cobj.get("scores", {}).get(k, 0.0)) for k in [c.name for c in criteria]}
            base = {
                "Файл": it["name"],
                "ФИО": fio,
                "Специализация": specialization,
                "Email": email_final,
                "Phone": phone_final,
                "FullText": text,
                "_FileHash": it["fh"], "_TextHash": it["th"],
                "_Reasoning": cobj.get("reasoning", {}),
            }
            for c in criteria:
                base[f"{c.name} (0-5)"] = score_map.get(c.name, 0.0)

            rows.append(base)
            texts.append(text); filenames.append(it["name"])

        # обновляем единый прогресс-бар по батчам
        progress_bar.progress(bi/num_batches)

    if not rows:
        st.warning("Нет успешных результатов после LLM"); st.stop()

    # -------- Pass 3: similarity + email/phone duplicate removal --------
    sim_max, sim_near, pairs_df = max_similarities(texts, filenames)
    for idx, row in enumerate(rows):
        row["SimilarityMax"] = round(sim_max[idx], 1)
        row["NearDuplicateOf"] = sim_near[idx]

    by_email: Dict[str,int] = {}
    by_phone: Dict[str,int] = {}
    keep_mask = [True]*len(rows)
    def norm_email(e: str) -> str: return (e or "").strip().lower()
    def norm_phone(p: str) -> str: return "".join(d for d in (p or "") if d.isdigit())

    for i,r in enumerate(rows):
        e = norm_email(r.get("Email",""))
        if e:
            if e in by_email: keep_mask[i] = False
            else: by_email[e] = i
    for i,r in enumerate(rows):
        if not keep_mask[i]: continue
        p = norm_phone(r.get("Phone",""))
        if p:
            if p in by_phone: keep_mask[i] = False
            else: by_phone[p] = i

    if not pairs_df.empty:
        name_to_idx = {rows[i]["Файл"]: i for i in range(len(rows))}
        for _, r in pairs_df.iterrows():
            a, b, s = r["FileA"], r["FileB"], float(r["Similarity"])
            ia, ib = name_to_idx.get(a), name_to_idx.get(b)
            if ia is None or ib is None: continue
            if not keep_mask[ia] or not keep_mask[ib]: continue
            if s >= 100.0 or s >= dup_threshold:
                drop = ib if ia < ib else ia
                keep_mask[drop] = False

    rows = [r for i,r in enumerate(rows) if keep_mask[i]]

    # -------- Таблицы и вывод --------
    df = compute_scores_table(rows, criteria, dup_threshold)

    show_cols = [
        "Файл","ФИО","Специализация","Email","Phone",
        "Coverage","CompositeScore","PriorityBucket","SimilarityMax","NearDuplicateOf","CalcComment"
    ]
    crit_cols = [f"{c.name} (0-5)" for c in criteria if f"{c.name} (0-5)" in df.columns]
    final_df = df[[c for c in show_cols if c in df.columns] + crit_cols].copy()
    final_df.insert(0, "Rank", range(1, len(final_df)+1))
    final_df = final_df.sort_values(["CompositeScore","SimilarityMax"], ascending=[False, False]).reset_index(drop=True)
    final_df["Rank"] = range(1, len(final_df)+1)

    # Листы статистики
    stats = []
    for c in criteria:
        pcol = f"{c.name}::Pct"
        if pcol in df:
            s = df[pcol].dropna()
            stats.append({
                "Criterion": c.name, "Weight": float(c.weight),
                "MedianPct": float(s.median()) if len(s) else 0.0,
                "P10": float(s.quantile(0.10)) if len(s) else 0.0,
                "P90": float(s.quantile(0.90)) if len(s) else 0.0,
                "CoverageShare": float((df[f"{c.name} (0-5)"]>0).mean()) if f"{c.name} (0-5)" in df else 0.0
            })
    crit_stats_df = pd.DataFrame(stats)
    similarity_pairs_df = pairs_df.head(200).copy() if add_pairs and not pairs_df.empty else pd.DataFrame(columns=["FileA","FileB","Similarity"])

    logic_text = (
        "Как считалось:\n"
        "• LLM выставляет баллы 0–5 по вашим критериям и даёт пояснения.\n"
        "• Композит = 0.75×взвешенные перцентили + 0.25×покрытие.\n"
        "• Дубликаты: одинаковые файлы/тексты, одинаковые email/телефоны удаляются; Similarity ≥ порога помечается как риск.\n"
        "• Комментарий содержит ФИО/специализацию, топ-3 сильных с выдержками, пробелы (низкие баллы) и риски."
    )
    config_df = pd.DataFrame({
        "Key":[
            "Model","JobTitle","DuplicateThreshold","CheckpointFile",
            "TotalUploaded","KeptAfterLocalDedup","BatchSize","NumBatches","HumanLogic"
        ],
        "Value":[
            model_name, job_title, str(dup_threshold), cp_path,
            str(len(files)), str(len(parsed_items)), "5", str(len(list(chunks(parsed_items, 5)))), logic_text
        ]
    })

    # Write XLSX with formatting
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.utils import get_column_letter

    DATA_BAR_COLOR = Color("FF63BE7B")  # зелёный ARGB

    xlsx_buf = io.BytesIO()
    with pd.ExcelWriter(xlsx_buf, engine="openpyxl") as xw:
        final_df.to_excel(xw, sheet_name="Ranking", index=False)
        crit_stats_df.to_excel(xw, sheet_name="CriteriaStats", index=False)
        similarity_pairs_df.to_excel(xw, sheet_name="SimilarityPairs", index=False)
        config_df.to_excel(xw, sheet_name="Config", index=False)

        wb = xw.book
        ws = wb["Ranking"]

        # Bold headers + borders
        thin = Side(border_style="thin", color="DDDDDD")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.border = border_all
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in r:
                cell.border = border_all
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Freeze header + autofilter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Zebra stripes (light)
        stripe = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
        for row in range(2, ws.max_row+1, 2):
            for col in range(1, ws.max_column+1):
                ws.cell(row=row, column=col).fill = stripe

        # Column widths
        for col_idx in range(1, ws.max_column+1):
            col = get_column_letter(col_idx)
            max_len = max(len(str(ws[f"{col}{r}"].value)) if ws[f"{col}{r}"].value is not None else 0 for r in range(1, ws.max_row+1))
            ws.column_dimensions[col].width = min(50, max(12, max_len + 2))

        headers = {cell.value: cell.col_idx for cell in ws[1]}
        def col_letter(name: str) -> str: return get_column_letter(headers[name])

        # CompositeScore color scale
        if "CompositeScore" in headers:
            c = col_letter("CompositeScore")
            ws.conditional_formatting.add(
                f"{c}2:{c}{ws.max_row}",
                ColorScaleRule(start_type="min", start_color="F8696B",
                               mid_type="percentile", mid_value=50, mid_color="FFEB84",
                               end_type="max", end_color="63BE7B")
            )

        # Data bars for criteria
        for h, idx in headers.items():
            if h.endswith(" (0-5)"):
                col = get_column_letter(idx)
                rule = DataBarRule(
                    start_type="num", start_value=0,
                    end_type="num", end_value=5,
                    color=DATA_BAR_COLOR,
                    showValue=True
                )
                ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", rule)

        # Row highlight by PriorityBucket
        if "PriorityBucket" in headers:
            bc = col_letter("PriorityBucket")
            for row in range(2, ws.max_row+1):
                v = str(ws[f"{bc}{row}"].value or "")
                fill = None
                if v == "A": fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
                elif v == "B": fill = PatternFill(start_color="FFF5CC", end_color="FFF5CC", fill_type="solid")
                elif v == "C": fill = PatternFill(start_color="FDE7E9", end_color="FDE7E9", fill_type="solid")
                if fill:
                    for col in range(1, ws.max_column+1):
                        ws.cell(row=row, column=col).fill = fill

        # Risk highlight: SimilarityMax >= dup_threshold
        if "SimilarityMax" in headers:
            sc = col_letter("SimilarityMax")
            for row in range(2, ws.max_row+1):
                try:
                    val = float(ws[f"{sc}{row}"].value)
                    if val >= dup_threshold:
                        fill = PatternFill(start_color="FFE8AA", end_color="FFE8AA", fill_type="solid")
                        for col in range(1, ws.max_column+1):
                            ws.cell(row=row, column=col).fill = fill
                        ws[f"{sc}{row}"].font = Font(bold=True)
                except Exception:
                    pass

    data = xlsx_buf.getvalue()

    # Save to server
    try:
        if save_path:
            with open(save_path, "wb") as f: f.write(data)
            st.success(f"Файл сохранён: {save_path}")
    except Exception as e:
        st.warning(f"Не удалось сохранить на сервере: {e}")

    # Download
    st.download_button("⬇️ Скачать XLSX", data=data,
                       file_name=os.path.basename(save_path) or "resume_ranking.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Footer branding
st.markdown("---")
st.caption("© shlukha, Liza & Partners LLC — 🗑️ clean_the_garbage.exe")
